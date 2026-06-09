# routes/dashboard.py
from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from models import db, MTCOrder
from models import db, MTCOrder, ConfigMacThep
import os
import base64
from openpyxl import load_workbook, Workbook
import io
import uuid
import pandas as pd
from auth.decorator import permission_required
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as XLImage
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from PIL import Image as PILImage
from models import db, MTCOrder, ConfigMacThep
from openpyxl.worksheet.page import PageMargins 
import math

mtc_bp = Blueprint('mtc_bp', __name__)

# Trang hiển thị danh sách
@mtc_bp.route('/mtc')
@permission_required('manage_mtc')
def mtc():
    # Lấy toàn bộ dữ liệu, sắp xếp dòng mới nhất lên đầu
    orders = MTCOrder.query.order_by(MTCOrder.ID.desc()).limit(1000).all()
    return render_template('mtc/mtc.html', orders=orders)

# API Đổi trạng thái (Process <-> Success)
@mtc_bp.route('/api/mtc/update-status/<int:id>', methods=['POST'])
@permission_required('manage_mtc')
def update_status(id):
    order = MTCOrder.query.get_or_404(id)
    data = request.json or {}
    new_status = data.get('new_status')
    
    # Đảm bảo tương thích ngược với các dòng dữ liệu cũ trong DB
    current_status = order.Trang_Thai or 'DANG_CHOT'
    if current_status == 'Process':
        current_status = 'DANG_CHOT'
    elif current_status == 'Success':
        current_status = 'CHOT_IN'

    # --- VALIDATOR LUỒNG TRẠNG THÁI NGHIỆM VỤ NGHIÊM NGẶT ---
    # Luồng A: Từ Đang chốt -> Tiến lên Chốt in (Khóa đơn + Gọi Iframe tạo file PNG)
    if current_status == 'DANG_CHOT' and new_status == 'CHOT_IN':
        action_required = "none"
        
    # Luồng B: Từ Chốt in -> Tiến lên Đã in xong (Xưởng báo cáo tiến độ)
    elif current_status == 'CHOT_IN' and new_status == 'DA_IN':
        action_required = "none"
        
    # Luồng C: Từ Đã in xong -> Nhảy cóc lùi về Chốt in (In lại nếu hỏng phôi, dữ liệu vẫn khóa)
    elif current_status == 'DA_IN' and new_status == 'CHOT_IN':
        action_required = "none"
        
    else:
        return jsonify({"status": "error", "message": "Thao tác chuyển đổi trạng thái không đúng luồng nghiệp vụ!"}), 400

    order.Trang_Thai = new_status
    db.session.commit()
    
    return jsonify({
        "status": "success", 
        "new_status": new_status, 
        "action_required": action_required
    })

# API Xóa dòng dữ liệu
@mtc_bp.route('/api/mtc/delete-order/<int:id>', methods=['POST'])
@permission_required('manage_mtc')
def delete_order(id):
    order = MTCOrder.query.get_or_404(id)
    
    # THÊM LỚP BẢO VỆ NÀY: Khóa tính năng xóa nếu đã chốt in
    current_status = order.Trang_Thai or 'DANG_CHOT'
    if current_status in ['CHOT_IN', 'DA_IN', 'Success']:
        return jsonify({"status": "error", "message": "Đơn hàng đã được chốt in hoặc in xong, không thể xóa!"}), 400
        
    db.session.delete(order)
    db.session.commit()
    return jsonify({"status": "success"})
@mtc_bp.route('/preview/mtc/<int:id>')
@permission_required('manage_mtc')
def preview_label(id):
    # Lấy thông tin đơn hàng từ DB
    order = MTCOrder.query.get_or_404(id)

    # 1. Xác định Class CSS (cert-sni, cert-ce, cert-ms-left, cert-ms-right)
    loai_mtc_lower = (order.Loai_MTC or "").lower()
    
    if 'sni' in loai_mtc_lower:
        cert_class = 'cert-sni'
    elif 'ce' in loai_mtc_lower:
        cert_class = 'cert-ce'
    elif 'ms' in loai_mtc_lower:
        cert_class = 'cert-ms-right'
    else:
        cert_class = 'cert-none'

    # 2. Tìm Tiêu chuẩn & License trong bảng Cấu hình (Config_MacThep)
    search_loai = 'MS' if 'MS' in order.Loai_MTC.upper() else order.Loai_MTC

    # 2. Mang search_loai đi tìm trong bảng Config
    config = ConfigMacThep.query.filter_by(Loai_MTC=search_loai, Mac_Thep=order.Mac_Thep).first()
    if not config:
        config = ConfigMacThep.query.filter_by(Loai_MTC=search_loai, Mac_Thep='DEFAULT').first()

    return render_template('mtc/mtc_preview.html', order=order, cert_class=cert_class, config=config)
# 1. API Hiển thị trang Edit
@mtc_bp.route('/edit/mtc/<int:id>', methods=['GET'])
@permission_required('manage_mtc')
def edit_page(id):
    order = MTCOrder.query.get_or_404(id)
    return render_template('mtc/mtc_edit.html', order=order)

# 2. API Xử lý Cập nhật dữ liệu
@mtc_bp.route('/api/update-order/mtc/<int:id>', methods=['POST'])
@permission_required('manage_mtc')
def update_order(id):
    order = MTCOrder.query.get_or_404(id)
    
    # Lớp kiểm tra bảo mật Backend (Chỉ cho sửa khi trạng thái là Đang chốt)
    current_status = order.Trang_Thai or 'DANG_CHOT'
    if current_status in ['CHOT_IN', 'DA_IN', 'Success']:
        return jsonify({"status": "error", "message": "Đơn hàng này đang ở trạng thái xử lý in, không cho phép chỉnh sửa!"}), 400
    
    so_number = request.form.get('so_number', '').strip()
    mac_thep = request.form.get('mac_thep', '').strip()
    noi_dung_chinh = request.form.get('noi_dung_chinh', '').strip()
    
    if not mac_thep:
        mac_thep = 'DEFAULT'
    if not so_number:
        so_number = ''
        
    order.SO_Number = so_number
    order.Mac_Thep = mac_thep
    order.Noi_Dung_Chinh = noi_dung_chinh
    order.nhan_mau = request.form.get('nhan_mau', '').strip()
    order.pic = request.form.get('pic', '').strip()
    order.ten_tau = request.form.get('ten_tau', '').strip()
    order.ghi_chu = request.form.get('ghi_chu', '').strip()
    db.session.commit()
    return jsonify({"status": "success", "message": "Cập nhật thành công!"})

# 3. API Nhận ảnh từ Web và Lưu xuống ổ cứng (Local)
@mtc_bp.route('/api/mtc/save-local-file', methods=['POST'])
@permission_required('manage_mtc')
def save_local_file():
    data = request.json
    img_data = data.get('image')
    order_id = data.get('order_id')
    
    order = MTCOrder.query.get(order_id)
    if not order:
        return jsonify({"status": "error", "message": "Không tìm thấy đơn hàng"}), 404

    random_str = uuid.uuid4().hex[:8]
    filename = f"MTC_{order.SO_Number}_{random_str}.png"
    EXPORT_DIR = os.path.join(os.getcwd(), 'static', 'MTC_Exports')
    if not os.path.exists(EXPORT_DIR):
        os.makedirs(EXPORT_DIR)
        
    file_path = os.path.join(EXPORT_DIR, filename)
    
    try:
        if "base64," in img_data:
            img_data = img_data.split("base64,")[1]
        with open(file_path, "wb") as fh:
            fh.write(base64.b64decode(img_data))
        
        # SỬA Ở ĐÂY: Thêm '/static/' vào đầu chuỗi link
        order.linkPng = f"/static/MTC_Exports/{filename}"
        db.session.commit()
        
        return jsonify({"status": "success", "message": "Đã lưu file và cập nhật DB!"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
@mtc_bp.route('/create/mtc', methods=['GET'])
@permission_required('manage_mtc')
def create_page():
    return render_template('mtc/mtc_create.html')

# 2. API Nhận dữ liệu tạo mới lưu vào DB
@mtc_bp.route('/api/mtc/create-order', methods=['POST'])
@permission_required('manage_mtc')
def create_order():
    try:
        loai_mtc = request.form.get('loai_mtc', '').strip()
        so_number = request.form.get('so_number', '').strip()
        mac_thep = request.form.get('mac_thep', '').strip()
        noi_dung_chinh = request.form.get('noi_dung_chinh', '').strip()
        nhan_mau = request.form.get('nhan_mau', '').strip()
        pic = request.form.get('pic', '').strip()
        ten_tau = request.form.get('ten_tau', '').strip()
        ghi_chu = request.form.get('ghi_chu', '').strip()
        # Xử lý thông minh: Nếu để trống mác thép thì tự động gán là 'DEFAULT'
        if not mac_thep:
            mac_thep = 'DEFAULT'
        if not so_number:
            so_number = ''    
        new_order = MTCOrder(
            Loai_MTC=loai_mtc,
            SO_Number=so_number,
            Mac_Thep=mac_thep,
            Noi_Dung_Chinh=noi_dung_chinh,
            nhan_mau=nhan_mau,
            pic=pic,
            ten_tau=ten_tau,
            ghi_chu=ghi_chu
        )
        
        db.session.add(new_order)
        db.session.commit()
        return jsonify({"status": "success", "message": "Tạo tem mới thành công!"})
    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Lỗi: {str(e)}"}), 500

from PIL import Image as PILImage, ImageDraw, ImageFont

from PIL import Image as PILImage, ImageDraw, ImageFont
import io
import os
from flask import send_file, current_app
from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins 
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

from PIL import Image as PILImage, ImageDraw, ImageFont
import io, os
from flask import send_file, current_app
from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins 
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

def get_best_fit_font_and_wrap(text, max_w_px, max_h_px, draw, font_path):
    sizes = [18.0, 17.0, 16.0, 15.0, 14.0, 13.5, 13.0]
    raw_lines = text.split('\n')
    
    for size in sizes:
        try: font = ImageFont.truetype(font_path, int(size))
        except: font = ImageFont.load_default()
        
        fits_width = True
        for line in raw_lines:
            if line.strip() and draw.textlength(line, font=font) > max_w_px:
                fits_width = False; break
        
        if fits_width and (len(raw_lines) * size * 1.3) <= max_h_px:
            return size, text 
            
    size = 13.0
    try: font = ImageFont.truetype(font_path, int(size))
    except: font = ImageFont.load_default()
    
    final_lines = []
    for line in raw_lines:
        if not line.strip():
            final_lines.append("")
            continue
        if draw.textlength(line, font=font) > max_w_px:
            words = line.split(' ')
            current_line = ""
            for word in words:
                test_line = f"{current_line} {word}".strip()
                if draw.textlength(test_line, font=font) <= max_w_px:
                    current_line = test_line
                else:
                    if current_line: final_lines.append(current_line)
                    current_line = word
            if current_line: final_lines.append(current_line)
        else:
            final_lines.append(line)
            
    return size, "\n".join(final_lines)

@mtc_bp.route('/api/mtc/export-excel/<int:id>', methods=['GET'])

@permission_required('manage_mtc')

def export_excel(id):

    order = MTCOrder.query.get_or_404(id)

    loai_raw = (order.Loai_MTC or "").upper()

   

    config = ConfigMacThep.query.filter_by(Loai_MTC=('MS' if 'MS' in loai_raw else loai_raw), Mac_Thep=order.Mac_Thep).first()

    if not config: config = ConfigMacThep.query.filter_by(Loai_MTC=('MS' if 'MS' in loai_raw else loai_raw), Mac_Thep='DEFAULT').first()



    std, lic = (config.Tieu_Chuan if config else ''), (config.License_No if config else '')

   

    # Xử lý triệt để biến mác thép (Xóa DEFAULT)f

    grade_val = order.Mac_Thep if order.Mac_Thep else ""

    mac_val = grade_val if grade_val != 'DEFAULT' else ""

    grade_for_dop = grade_val if grade_val and grade_val != 'DEFAULT' else 'S355JR'



    wb = Workbook(); ws = wb.active; ws.title = "ETIKET_FINAL"

    ws.page_margins = PageMargins(left=0, right=0, top=0.1, bottom=0.393, header=0, footer=0)

    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1

    # 2. Căn giữa tuyệt đối trên mặt giấy in (cả chiều ngang và dọc)
    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False

    # 3. KHÓA VÙNG IN (Cực kỳ quan trọng): 
    ws.print_area = 'A1:B6'


    ws.row_dimensions[1].height = 120.0; ws.row_dimensions[2].height = 9.75

    for i in range(3, 7): ws.row_dimensions[i].height = 34.5



    cert_type = 'sni' if 'SNI' in loai_raw else ('ce' if 'CE' in loai_raw else ('ms' if 'MS' in loai_raw else 'none'))
    if cert_type in ['ms', 'none'] and mac_val:
        # Dành cho MS khi có mác thép
        ws.column_dimensions['A'].width = 7.5
        ws.column_dimensions['B'].width = 52.43
        max_width_px = 280 # Thu hẹp khung text khoảng 6-7px so với bình thường
    else:
        # Dành cho các loại khác (CE, SNI, NONE)
        ws.column_dimensions['A'].width = 6.29
        ws.column_dimensions['B'].width = 53.14
        max_width_px = 295
    # CỘT A: Sửa lỗi hiển thị chữ DEFAULT

    ws.merge_cells('A3:A6')

    if cert_type in ['ms', 'none']:

        ws['A3'] = f"{order.SO_Number}\n{mac_val}" if mac_val else order.SO_Number

    else:

        ws['A3'] = order.SO_Number

       

    ws['A3'].alignment = Alignment(text_rotation=180, horizontal='right', vertical='center', wrapText=True)

    if cert_type in ['ms', 'none'] and mac_val:
        # Khi có 2 dòng dọc (Có mác thép): MS xài size 13, NONE xài size 16 (theo ý tưởng của bạn)
        a3_size = 14 
    else:
        # Khi chỉ có 1 dòng dọc (Không mác thép): Giữ size 18 cho to và rõ
        a3_size = 18

    ws['A3'].font = Font(name='Arial', bold=True, italic=True, size=a3_size)



    # NỘI DUNG CHÍNH (Đã hoạt động tốt)

    font_path = "C:\\Windows\\Fonts\\arialbd.ttf"

    temp_draw = ImageDraw.Draw(PILImage.new('RGBA', (1, 1)))

    raw_content = (order.Noi_Dung_Chinh or "").strip()

    raw_content = raw_content.replace('\r\n', '\n')

    best_size, final_text = get_best_fit_font_and_wrap(raw_content,max_width_px, 170, temp_draw, font_path)

   

    ws.merge_cells('B3:B6')

    ws['B3'] = final_text

    ws['B3'].font = Font(name='Arial', bold=True, size=best_size)

    ws['B3'].alignment = Alignment(wrapText=True, horizontal='center', vertical='center')



    # LOGO

    if cert_type != 'none':

        logo_p = os.path.join(current_app.root_path, 'static', 'mtc', f"logo_{cert_type}.png")

        if os.path.exists(logo_p):

            t_size = 145; canvas = PILImage.new('RGBA', (t_size, t_size), (255, 255, 255, 0)); draw = ImageDraw.Draw(canvas)

            logo_img = PILImage.open(logo_p).convert("RGBA")

            if cert_type == 'ce':
                max_height = int(t_size * 0.5)  # CE dẹt ngang, ép chiều cao 50% là vừa đẹp
            elif cert_type == 'ms':
                max_height = int(t_size * 0.62) # MS hình thoi, nới lỏng chiều cao lên 65% để hình to hơn
            else:
                max_height = t_size

            # Áp dụng resize
            logo_img.thumbnail((t_size, max_height), PILImage.LANCZOS)

            canvas.paste(logo_img, ((t_size - logo_img.width) // 2, 0), logo_img)

           

            try:

                f_14b = ImageFont.truetype(font_path, 14)

                f_11b = ImageFont.truetype(font_path, 11)

                f_9b = ImageFont.truetype(font_path, 9)  

            except:

                f_14b = f_11b = f_9b = ImageFont.load_default()



            if cert_type == 'ce':

                y = logo_img.height + 5

                draw.text(((t_size - draw.textlength(lic or "2195", f_14b)) // 2, y), lic or "2195", font=f_14b, fill=(0,0,0))

                y += 10 + 14; draw.text(((t_size - draw.textlength(std or "EN 10025-2", f_9b)) // 2, y), std or "EN 10025-2", font=f_9b, fill=(0,0,0))

                y += 5 + 9; draw.text(((t_size - draw.textlength(f"DoP: HP{grade_for_dop}AR", f_9b)) // 2, y), f"DoP: HP{grade_for_dop}AR", font=f_9b, fill=(0,0,0))

            elif cert_type == 'ms':

                y = logo_img.height + 5

                for txt in ["SIRIM", std or "", lic or ""]:

                    draw.text(((t_size - draw.textlength(txt, f_11b)) // 2, y), txt, font=f_11b, fill=(0,0,0))

                    y += 5 + 11



            img_s = io.BytesIO(); canvas.save(img_s, format='PNG'); img_s.seek(0)

            xl_img = XLImage(img_s); xl_img.width = xl_img.height = t_size

           

            # SỬA LỖI LOGO: Đặt rowOff = 5 (cách đỉnh ô 5px an toàn), tránh tọa độ âm làm mất hình!

            marker = AnchorMarker(col=1, colOff=pixels_to_EMU(int(372 - t_size - 12)), row=0, rowOff=pixels_to_EMU(5))

            xl_img.anchor = OneCellAnchor(_from=marker, ext=XDRPositiveSize2D(pixels_to_EMU(t_size), pixels_to_EMU(t_size)))

            ws.add_image(xl_img)



    out = io.BytesIO(); wb.save(out); out.seek(0)

    return send_file(out, as_attachment=True, download_name=f"ETIKET_{order.SO_Number}.xlsx")
@mtc_bp.route('/api/mtc/upload-excel', methods=['POST'])
@permission_required('manage_mtc')
def process_excel():
    if 'file' not in request.files:
        return jsonify({"status": "error", "message": "Không tìm thấy file gửi lên!"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"status": "error", "message": "Bạn chưa chọn file nào!"}), 400

    try:
        # Đọc Excel bằng Pandas
        df = pd.read_excel(file)
        
        df = df.iloc[::-1]
        # Làm sạch dữ liệu: Xóa các khoảng trắng thừa ở tên cột
        df.columns = df.columns.str.strip()
        
        # 1. Bắt buộc form Excel phải có đủ 4 cột này
        required_cols = ['Loai_MTC', 'SO_Number', 'Mac_Thep', 'Noi_Dung_Chinh']
        for col in required_cols:
            if col not in df.columns:
                return jsonify({"status": "error", "message": f"File Excel sai biểu mẫu. Thiếu cột: {col}"}), 400

        # Mẹo: Thay thế toàn bộ giá trị trống (NaN của Pandas) thành chuỗi rỗng '' để dễ so sánh
        df = df.fillna('')

        # Tạo một mảng tạm để chứa các đơn hàng nếu qua được vòng kiểm duyệt
        valid_orders = []

        # 2. Duyệt qua từng dòng và kiểm tra khắt khe
        for index, row in df.iterrows():
            row_excel = index + 2  # Tính chính xác số dòng trong file Excel (bao gồm cả dòng tiêu đề)
            
            # Lấy dữ liệu và làm sạch khoảng trắng
            loai_mtc_raw = str(row['Loai_MTC']).strip()
            so_number = str(row['SO_Number']).strip()
            mac_thep = str(row['Mac_Thep']).strip()
            noi_dung = str(row['Noi_Dung_Chinh']).strip()
            nhan_mau_raw = str(row.get('Nhan_mau', '')).strip() if pd.notna(row.get('Nhan_mau')) else ''
            pic_raw = str(row.get('PIC', '')).strip() if pd.notna(row.get('PIC')) else ''
            ten_tau_raw = str(row.get('TEN_TAU', '')).strip() if pd.notna(row.get('TEN_TAU')) else ''
            ghi_chu_raw = str(row.get('Ghi_chu', '')).strip() if pd.notna(row.get('Ghi_chu')) else ''
            # --- CHUẨN HÓA DỮ LIỆU ---
            loai_mtc = loai_mtc_raw.upper() if loai_mtc_raw else 'NONE'
            mac_thep_save = mac_thep if mac_thep else 'DEFAULT'
            so_number_save = so_number if so_number else ''

            # --- LOGIC VALIDATE TỪNG LOẠI MTC ---
            if loai_mtc == 'NONE':
                # NONE: Bắt buộc phải có Nội dung chính
                if not so_number_save:
                    return jsonify({"status": "error", "message": f"Lỗi dòng {row_excel}: Mẫu 'NONE' bắt buộc bắt buộc phải có Mã SO!"}), 400

            elif 'MS' in loai_mtc:
                # MS: Bắt buộc  + SO + Mác thép
                if not so_number_save or mac_thep_save == 'DEFAULT':
                    return jsonify({"status": "error", "message": f"Lỗi dòng {row_excel}: Mẫu '{loai_mtc_raw}' bắt buộc phải điền đủ Mã SO, Mác Thép !"}), 400

            elif 'CE' in loai_mtc or 'SNI' in loai_mtc:
                # CE và SNI: Bắt buộc SO + Nội dung chính (Mác thép cho phép trống -> tự động thành DEFAULT)
                if  not so_number_save:
                    return jsonify({"status": "error", "message": f"Lỗi dòng {row_excel}: Mẫu '{loai_mtc_raw}' bắt buộc phải có Mã SO!"}), 400
            
            else:
                # Nếu người dùng tự nhập một loại chứng chỉ lạ nào đó -> Cư xử như NONE
                if not so_number_save:
                    return jsonify({"status": "error", "message": f"Lỗi dòng {row_excel}: Loại MTC '{loai_mtc_raw}' bắt buộc phải có Mã SO!"}), 400

            # 3. Nếu dòng dữ liệu hợp lệ, đưa vào mảng chờ
            valid_orders.append(MTCOrder(
                Loai_MTC = loai_mtc_raw if loai_mtc_raw else 'NONE',
                SO_Number = so_number_save,
                Mac_Thep = mac_thep_save,
                Noi_Dung_Chinh = noi_dung,
                nhan_mau = nhan_mau_raw,
                pic = pic_raw,
                ten_tau = ten_tau_raw,
                ghi_chu = ghi_chu_raw
            ))

        # 4. Lưu toàn bộ xuống Database (Chỉ chạy đến đây khi KHÔNG có lỗi nào xảy ra)
        if valid_orders:
            # Dùng bulk_save_objects để insert cực nhanh hàng ngàn dòng cùng lúc
            db.session.bulk_save_objects(valid_orders)
            db.session.commit()
            return jsonify({"status": "success", "message": f"Đã Validate và lưu thành công {len(valid_orders)} đơn hàng!"}), 200
        else:
            return jsonify({"status": "error", "message": "File Excel trống, không có dữ liệu để lưu."}), 400

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": f"Lỗi trong quá trình xử lý: {str(e)}"}), 500